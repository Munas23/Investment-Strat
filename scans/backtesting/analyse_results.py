"""
Backtest Performance Analyser

Loads the latest results for all 5 systems and generates:
  1. Excel workbook         — Summary + per-system trade logs + monthly returns table
  2. equity_curves.png      — All systems vs SPY equity growth
  3. drawdowns.png          — Underwater drawdown chart per system
  4. monthly_returns.png    — Calendar heatmap of monthly returns
  5. pnl_distribution.png   — Trade P&L histogram per system
  6. annual_returns.png     — Grouped annual returns bar chart
  7. trade_scatter.png      — Holding days vs P&L % scatter
  8. exit_reasons.png       — Pie chart of exit reasons per system

Usage:
    python analyse_results.py
    python analyse_results.py --results-dir path/to/results
    python analyse_results.py --output-dir path/to/output
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use('Agg')   # Non-interactive backend — works without a display
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd

try:
    import seaborn as sns
    HAS_SEABORN = True
except ImportError:
    HAS_SEABORN = False

try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SYSTEM_PREFIXES = {
    1: 'Conservative_Growth_(Minervini)',
    2: 'Turtle_ATR-Based',
    3: 'Qullamaggie_Aggressive',
    4: 'Hybrid_Balanced',
    5: 'High_Conviction_Only',
}

DISPLAY_NAMES = {
    1: 'Sys1 Minervini',
    2: 'Sys2 Turtle',
    3: 'Sys3 Qullamaggie',
    4: 'Sys4 Hybrid',
    5: 'Sys5 High Conv',
}

COLORS = {
    1: '#2196F3',
    2: '#4CAF50',
    3: '#FF5722',
    4: '#9C27B0',
    5: '#FF9800',
    'SPY': '#607D8B',
}

ACCOUNT_SIZE    = 2_000_000
BENCHMARK_CAGR  = 14.33    # percent — matches comparison.csv

MONTHS_SHORT = ['Jan','Feb','Mar','Apr','May','Jun',
                'Jul','Aug','Sep','Oct','Nov','Dec']


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def find_latest_files(results_dir: Path) -> dict:
    """
    For each system find the most recent trades / equity / metrics triple.
    Returns {system_id: {'trades': Path, 'equity': Path, 'metrics': Path, 'name': str}}
    """
    found = {}
    for sys_id, prefix in SYSTEM_PREFIXES.items():
        # Sort alphabetically = chronologically because timestamp is ISO-like
        trades_files = sorted(results_dir.glob(f"{prefix}_*_trades.csv"))
        if not trades_files:
            print(f"  [skip] No trades file for System {sys_id} ({prefix})")
            continue

        latest = trades_files[-1]
        stem   = latest.stem.replace('_trades', '')

        equity  = results_dir / f"{stem}_equity.csv"
        metrics = results_dir / f"{stem}_metrics.json"

        if not equity.exists() or not metrics.exists():
            print(f"  [skip] Missing equity/metrics for System {sys_id}")
            continue

        found[sys_id] = {
            'trades':  latest,
            'equity':  equity,
            'metrics': metrics,
            'name':    DISPLAY_NAMES[sys_id],
        }
        print(f"  Sys{sys_id}: {latest.name}")

    return found


# ---------------------------------------------------------------------------
# Data loading helpers
# ---------------------------------------------------------------------------

def load_system_data(info: dict) -> dict:
    trades  = pd.read_csv(info['trades'],  parse_dates=['entry_date', 'exit_date'])
    equity  = pd.read_csv(info['equity'],  parse_dates=['date'], index_col='date')
    with open(info['metrics']) as f:
        metrics = json.load(f)
    return {'trades': trades, 'equity': equity, 'metrics': metrics}


def spy_equity_curve(index: pd.DatetimeIndex) -> pd.Series:
    """Reconstruct SPY buy-and-hold equity from benchmark CAGR."""
    days   = (index - index[0]).days
    values = ACCOUNT_SIZE * (1 + BENCHMARK_CAGR / 100) ** (days / 365)
    return pd.Series(values, index=index, name='SPY')


def drawdown_series(equity: pd.Series) -> pd.Series:
    """Percentage drawdown from rolling peak (0 to -100)."""
    peak = equity.cummax()
    return (equity - peak) / peak * 100


def monthly_return_table(equity: pd.Series) -> pd.DataFrame:
    """
    Return a year × month DataFrame of monthly returns (%).
    Columns are integers 1–12; call .rename(columns=...) for labels.
    """
    monthly = equity.resample('ME').last().pct_change().dropna() * 100
    monthly.index = monthly.index.to_period('M')
    tbl = monthly.groupby([monthly.index.year, monthly.index.month]).first().unstack(level=1)
    # unstack may produce a MultiIndex (value, month) — flatten to just month integers
    if isinstance(tbl.columns, pd.MultiIndex):
        tbl.columns = tbl.columns.get_level_values(-1)
    return tbl


def annual_returns(equity: pd.Series) -> pd.Series:
    """Annual returns (%) indexed by integer year."""
    ann = equity.resample('YE').last().pct_change().dropna() * 100
    ann.index = ann.index.year
    return ann


# ---------------------------------------------------------------------------
# Chart helpers
# ---------------------------------------------------------------------------

def _save(fig: plt.Figure, path: Path):
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {path.name}")


# Chart 1 — Equity curves
def plot_equity_curves(systems: dict, out: Path):
    fig, ax = plt.subplots(figsize=(14, 7))

    ref_idx = list(systems.values())[0]['equity']['equity'].index
    spy = spy_equity_curve(ref_idx)
    ax.plot(spy.index, spy / 1e6, '--', color=COLORS['SPY'],
            linewidth=1.5, label=f'SPY ({BENCHMARK_CAGR:.1f}% CAGR)', alpha=0.8)

    for sid, data in systems.items():
        eq   = data['equity']['equity']
        cagr = data['metrics']['cagr']
        ax.plot(eq.index, eq / 1e6, color=COLORS[sid], linewidth=2,
                label=f"{DISPLAY_NAMES[sid]} ({cagr:.1f}%)")

    ax.set_title('Equity Curves — All Systems vs SPY  (2020–2024)',
                 fontsize=14, fontweight='bold')
    ax.set_ylabel('Portfolio Value')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:.1f}M'))
    ax.legend(loc='upper left', fontsize=9)
    ax.grid(True, alpha=0.3)
    ax.set_facecolor('#fafafa')
    fig.tight_layout()
    _save(fig, out)


# Chart 2 — Drawdowns (stacked subplots)
def plot_drawdowns(systems: dict, out: Path):
    n   = len(systems)
    fig, axes = plt.subplots(n, 1, figsize=(14, 2.8 * n), sharex=True)
    if n == 1:
        axes = [axes]

    for ax, (sid, data) in zip(axes, systems.items()):
        eq = data['equity']['equity']
        dd = drawdown_series(eq)
        ax.fill_between(dd.index, dd.values, 0, color=COLORS[sid], alpha=0.4)
        ax.plot(dd.index, dd.values, color=COLORS[sid], linewidth=0.8)
        ax.axhline(0, color='black', linewidth=0.5)
        max_dd = data['metrics']['max_drawdown_pct']
        ax.set_title(f"{DISPLAY_NAMES[sid]}  |  Max DD: {max_dd:.1f}%", fontsize=10)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:.0f}%'))
        ax.grid(True, alpha=0.2)

    fig.suptitle('Drawdown — All Systems', fontsize=13, fontweight='bold')
    fig.tight_layout()
    _save(fig, out)


# Chart 3 — Monthly returns heatmap
def plot_monthly_returns(systems: dict, out: Path):
    n   = len(systems)
    fig, axes = plt.subplots(n, 1, figsize=(13, 3.2 * n))
    if n == 1:
        axes = [axes]

    VMAX = 10  # colour scale ±10%

    for ax, (sid, data) in zip(axes, systems.items()):
        tbl = monthly_return_table(data['equity']['equity'])
        # Fill missing months with NaN
        for m in range(1, 13):
            if m not in tbl.columns:
                tbl[m] = np.nan
        tbl = tbl[[m for m in range(1, 13)]]
        tbl.columns = MONTHS_SHORT

        if HAS_SEABORN:
            sns.heatmap(tbl, ax=ax, cmap='RdYlGn', center=0,
                        vmin=-VMAX, vmax=VMAX,
                        annot=True, fmt='.1f', annot_kws={'size': 8},
                        linewidths=0.5, cbar_kws={'label': '%'})
        else:
            im = ax.imshow(tbl.values, cmap='RdYlGn', vmin=-VMAX, vmax=VMAX, aspect='auto')
            ax.set_xticks(range(12))
            ax.set_xticklabels(MONTHS_SHORT, fontsize=8)
            ax.set_yticks(range(len(tbl.index)))
            ax.set_yticklabels(tbl.index, fontsize=8)
            for i in range(len(tbl.index)):
                for j in range(12):
                    v = tbl.values[i, j]
                    if not np.isnan(v):
                        ax.text(j, i, f'{v:.1f}', ha='center', va='center', fontsize=7)
            fig.colorbar(im, ax=ax, label='%')

        ax.set_title(f"{DISPLAY_NAMES[sid]}  Monthly Returns (%)",
                     fontsize=10, fontweight='bold')

    fig.suptitle('Monthly Returns Heatmap — All Systems', fontsize=13, fontweight='bold')
    fig.tight_layout()
    _save(fig, out)


# Chart 4 — P&L distribution histogram
def plot_pnl_distribution(systems: dict, out: Path):
    n   = len(systems)
    fig, axes = plt.subplots(n, 1, figsize=(12, 3 * n))
    if n == 1:
        axes = [axes]

    for ax, (sid, data) in zip(axes, systems.items()):
        pnl    = data['trades']['pnl_percent'].dropna()
        wins   = pnl[pnl > 0]
        losses = pnl[pnl <= 0]

        lo, hi = max(pnl.min(), -30), min(pnl.max(), 60)
        bins   = np.linspace(lo, hi, 50)

        ax.hist(losses, bins=bins, color='#ef5350', alpha=0.7, label=f'Losses ({len(losses)})')
        ax.hist(wins,   bins=bins, color='#66bb6a', alpha=0.7, label=f'Wins ({len(wins)})')
        ax.axvline(0,         color='black', linewidth=1.0)
        ax.axvline(pnl.mean(), color='navy', linewidth=1.5, linestyle='--',
                   label=f'Mean {pnl.mean():.1f}%')

        m = data['metrics']
        ax.set_title(
            f"{DISPLAY_NAMES[sid]}  |  Win Rate: {m['win_rate']:.1f}%  "
            f"|  Profit Factor: {m['profit_factor']:.2f}  "
            f"|  Avg Win: {m['avg_win_pct']:.1f}%  Avg Loss: {m['avg_loss_pct']:.1f}%",
            fontsize=9)
        ax.set_xlabel('P&L %')
        ax.set_ylabel('Trades')
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

    fig.suptitle('Trade P&L Distribution — All Systems', fontsize=13, fontweight='bold')
    fig.tight_layout()
    _save(fig, out)


# Chart 5 — Annual returns grouped bar
def plot_annual_returns(systems: dict, out: Path):
    ann_data = {}
    for sid, data in systems.items():
        eq = data['equity']['equity']
        ann_data[DISPLAY_NAMES[sid]] = annual_returns(eq)

    df = pd.DataFrame(ann_data)

    # Add SPY as flat benchmark reference line (not a bar — keep it clean)
    years    = df.index.tolist()
    n_sys    = len(df.columns)
    x        = np.arange(len(years))
    bar_w    = 0.7 / n_sys

    fig, ax = plt.subplots(figsize=(14, 6))

    for i, col in enumerate(df.columns):
        sid   = {v: k for k, v in DISPLAY_NAMES.items()}.get(col, 0)
        color = COLORS.get(sid, '#aaa')
        offset = (i - n_sys / 2 + 0.5) * bar_w
        ax.bar(x + offset, df[col].fillna(0), bar_w,
               label=col, color=color, alpha=0.85)

    # SPY horizontal dashed line
    ax.axhline(BENCHMARK_CAGR, color=COLORS['SPY'], linewidth=1.5,
               linestyle='--', label=f'SPY {BENCHMARK_CAGR:.1f}% avg')
    ax.axhline(0, color='black', linewidth=0.8)

    ax.set_xticks(x)
    ax.set_xticklabels(years)
    ax.set_ylabel('Annual Return %')
    ax.set_title('Annual Returns — All Systems vs SPY', fontsize=13, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, axis='y', alpha=0.3)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{v:.0f}%'))

    fig.tight_layout()
    _save(fig, out)


# Chart 6 — Holding days vs P&L scatter
def plot_trade_scatter(systems: dict, out: Path):
    n   = len(systems)
    fig, axes = plt.subplots(1, n, figsize=(4 * n, 5), sharey=True)
    if n == 1:
        axes = [axes]

    for ax, (sid, data) in zip(axes, systems.items()):
        trades = data['trades']
        pnl    = trades['pnl_percent'].clip(-25, 55)
        days   = trades['holding_days'].clip(0, 60)
        colors = ['#66bb6a' if v > 0 else '#ef5350' for v in pnl]

        ax.scatter(days, pnl, c=colors, alpha=0.25, s=12, edgecolors='none')
        ax.axhline(0, color='black', linewidth=0.8)
        ax.set_xlabel('Holding Days')
        ax.set_title(DISPLAY_NAMES[sid], fontsize=9)
        ax.grid(True, alpha=0.2)

    axes[0].set_ylabel('P&L %')
    fig.suptitle('Holding Duration vs P&L — All Systems', fontsize=13, fontweight='bold')
    fig.tight_layout()
    _save(fig, out)


# Chart 7 — Rolling Sharpe (252-day)
def plot_rolling_sharpe(systems: dict, out: Path, window: int = 252):
    fig, ax = plt.subplots(figsize=(14, 6))
    RISK_FREE = 0.04 / 252  # daily risk-free rate (4% annual)

    for sid, data in systems.items():
        eq      = data['equity']['equity']
        ret     = eq.pct_change().dropna()
        excess  = ret - RISK_FREE
        roll_sr = (excess.rolling(window).mean() / ret.rolling(window).std()) * np.sqrt(252)
        roll_sr = roll_sr.dropna()
        ax.plot(roll_sr.index, roll_sr.values, color=COLORS[sid],
                linewidth=1.5, label=DISPLAY_NAMES[sid])

    ax.axhline(1.0, color='grey', linewidth=0.8, linestyle='--', alpha=0.7, label='Sharpe = 1')
    ax.axhline(0,   color='black', linewidth=0.5)
    ax.set_title(f'Rolling Sharpe Ratio ({window}-day)  — Stability Check',
                 fontsize=13, fontweight='bold')
    ax.set_ylabel('Sharpe Ratio')
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3)
    ax.set_facecolor('#fafafa')
    fig.tight_layout()
    _save(fig, out)


# Chart 8 — Exit reasons pie
def plot_exit_reasons(systems: dict, out: Path):
    n   = len(systems)
    fig, axes = plt.subplots(1, n, figsize=(3.5 * n, 5))
    if n == 1:
        axes = [axes]

    PIE_COLORS = ['#66bb6a','#81c784','#a5d6a7',
                  '#ef5350','#e57373','#ef9a9a',
                  '#607D8B','#bdbdbd']

    for ax, (sid, data) in zip(axes, systems.items()):
        counts = data['trades']['exit_reason'].value_counts()
        ax.pie(counts.values, labels=counts.index,
               colors=PIE_COLORS[:len(counts)],
               autopct='%1.0f%%', startangle=90,
               textprops={'fontsize': 7})
        ax.set_title(DISPLAY_NAMES[sid], fontsize=9)

    fig.suptitle('Exit Reasons — All Systems', fontsize=13, fontweight='bold')
    fig.tight_layout()
    _save(fig, out)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def _col_width(df: pd.DataFrame, col: str) -> int:
    max_data = df[col].astype(str).str.len().max() if len(df) else 0
    return max(len(str(col)), int(max_data), 8) + 2


def build_excel(systems: dict, results_dir: Path, out: Path):
    if not HAS_OPENPYXL:
        print("  [skip] openpyxl not installed — run: pip install openpyxl")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import ColorScaleRule

    GREEN  = PatternFill('solid', fgColor='C8E6C9')
    RED    = PatternFill('solid', fgColor='FFCDD2')
    HEADER = PatternFill('solid', fgColor='1565C0')
    BOLD   = Font(bold=True)
    WHITE  = Font(bold=True, color='FFFFFF')
    CENTER = Alignment(horizontal='center')

    with pd.ExcelWriter(out, engine='openpyxl') as writer:

        # ── Sheet 1: Summary ──────────────────────────────────────────────
        comp_path = results_dir / 'comparison.csv'
        if comp_path.exists():
            summary = pd.read_csv(comp_path)
        else:
            rows = []
            for sid, data in systems.items():
                m = data['metrics']
                rows.append({
                    'System':        DISPLAY_NAMES[sid],
                    'CAGR':          m['cagr'],
                    'Alpha':         round(m['cagr'] - BENCHMARK_CAGR, 2),
                    'Max DD':        m['max_drawdown_pct'],
                    'Sharpe':        m['sharpe_ratio'],
                    'Sortino':       m['sortino_ratio'],
                    'Win Rate':      m['win_rate'],
                    'Profit Factor': m['profit_factor'],
                    'Avg R':         m['avg_r_multiple'],
                    'Trades':        m['total_trades'],
                    'Avg Hold Days': m['avg_holding_days'],
                    'Max Consec Loss': m.get('max_consecutive_losses', ''),
                })
            summary = pd.DataFrame(rows)

        summary.to_excel(writer, sheet_name='Summary', index=False)
        ws = writer.sheets['Summary']

        # Style header row
        for cell in ws[1]:
            cell.font    = WHITE
            cell.fill   = HEADER
            cell.alignment = CENTER

        # Auto column widths
        for col_cells in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = max_len + 3

        # Conditional colour on Alpha column
        alpha_col = None
        for i, h in enumerate(summary.columns, start=1):
            if str(h).lower() == 'alpha':
                alpha_col = i
                break
        if alpha_col:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(alpha_col)
            data_rows  = len(summary) + 1
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{data_rows}',
                ColorScaleRule(
                    start_type='min', start_color='FFCDD2',
                    mid_type='num',   mid_value=0, mid_color='FFFFFF',
                    end_type='max',   end_color='C8E6C9'
                )
            )

        # ── Per-system trade sheets ───────────────────────────────────────
        for sid, data in systems.items():
            sheet_name = f"Sys{sid}_{DISPLAY_NAMES[sid].split()[1]}"[:31]
            trades = data['trades'].copy()

            # Derived columns
            trades['win_loss'] = trades['pnl_percent'].apply(
                lambda x: 'WIN' if x > 0 else 'LOSS')

            # Sort by exit_date for readability
            if 'exit_date' in trades.columns:
                trades = trades.sort_values('exit_date')

            trades.to_excel(writer, sheet_name=sheet_name, index=False)
            ws_t = writer.sheets[sheet_name]

            # Header styling
            for cell in ws_t[1]:
                cell.font  = WHITE
                cell.fill  = HEADER
                cell.alignment = CENTER

            # Colour win/loss rows
            win_loss_col = None
            for i, h in enumerate(trades.columns, start=1):
                if h == 'win_loss':
                    win_loss_col = i
                    break

            if win_loss_col:
                from openpyxl.utils import get_column_letter
                wl_letter = get_column_letter(win_loss_col)
                for row_idx in range(2, len(trades) + 2):
                    cell = ws_t[f'{wl_letter}{row_idx}']
                    if cell.value == 'WIN':
                        cell.fill = GREEN
                    elif cell.value == 'LOSS':
                        cell.fill = RED

            # Auto column widths (sample first 200 rows for speed)
            for col_cells in ws_t.columns:
                sample = [c.value for c in col_cells[:200] if c.value is not None]
                max_len = max((len(str(v)) for v in sample), default=6)
                ws_t.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 25)

            # Freeze top row
            ws_t.freeze_panes = 'A2'

        # ── Monthly returns sheet ─────────────────────────────────────────
        row_cursor = 0
        first_sheet_written = False
        for sid, data in systems.items():
            tbl = monthly_return_table(data['equity']['equity'])
            for m in range(1, 13):
                if m not in tbl.columns:
                    tbl[m] = np.nan
            tbl = tbl[[m for m in range(1, 13)]]
            tbl.columns = MONTHS_SHORT
            tbl.index.name = 'Year'
            tbl = tbl.reset_index()

            # Write system label then table
            label_df = pd.DataFrame(
                {c: [f'=== {DISPLAY_NAMES[sid]} ===' if c == 'Year' else '']
                 for c in tbl.columns})
            block = pd.concat([label_df, tbl], ignore_index=True)

            block.to_excel(writer, sheet_name='Monthly_Returns',
                           index=False,
                           startrow=row_cursor,
                           header=(row_cursor == 0))
            row_cursor += len(block) + 2

        # Colour-scale the monthly returns data
        ws_m = writer.sheets['Monthly_Returns']
        ws_m.conditional_formatting.add(
            f'B2:{chr(ord("B") + 11)}{row_cursor}',
            ColorScaleRule(
                start_type='min', start_color='FFCDD2',
                mid_type='num',   mid_value=0, mid_color='FFFFFF',
                end_type='max',   end_color='C8E6C9'
            )
        )
        for cell in ws_m[1]:
            cell.font = WHITE
            cell.fill = HEADER

        # ── Equity curve data sheet ───────────────────────────────────────
        eq_df = pd.DataFrame()
        for sid, data in systems.items():
            s = data['equity']['equity'].rename(DISPLAY_NAMES[sid])
            if eq_df.empty:
                eq_df = s.to_frame()
                eq_df['SPY'] = spy_equity_curve(s.index)
            else:
                eq_df[DISPLAY_NAMES[sid]] = s

        eq_df.to_excel(writer, sheet_name='Equity_Data')

    print(f"  Saved: {out.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Analyse backtest results')
    parser.add_argument('--results-dir', type=Path, default=None,
                        help='Path to results directory (auto-detected if omitted)')
    parser.add_argument('--output-dir',  type=Path, default=None,
                        help='Where to write analysis files (default: results/../analysis)')
    args = parser.parse_args()

    script_dir = Path(__file__).parent

    # Locate results
    if args.results_dir:
        results_dir = args.results_dir
    else:
        # run_backtest.py saves relative to cwd, so results end up in
        # backtesting/backtesting/results/ when run from scans/backtesting/
        candidates = [
            script_dir / 'backtesting' / 'results',
            script_dir / 'results',
        ]
        results_dir = next((p for p in candidates if p.exists()), None)

    if results_dir is None or not results_dir.exists():
        print("ERROR: Could not find results directory.")
        print("Run run_backtest.py first, then re-run this script.")
        sys.exit(1)

    # Output directory
    if args.output_dir:
        output_dir = args.output_dir
    else:
        output_dir = results_dir.parent / 'analysis'
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')

    print(f"\nBacktest Performance Analyser")
    print(f"{'Results':12}: {results_dir}")
    print(f"{'Output':12}: {output_dir}")
    print(f"\nLocating latest run per system...")
    file_info = find_latest_files(results_dir)
    if not file_info:
        print("ERROR: No complete result sets found (need _trades + _equity + _metrics).")
        sys.exit(1)

    print(f"\nLoading data...")
    systems = {}
    for sid, info in file_info.items():
        systems[sid] = load_system_data(info)
        m = systems[sid]['metrics']
        print(f"  {DISPLAY_NAMES[sid]:<20}  {m['total_trades']:>5} trades  "
              f"CAGR {m['cagr']:>6.2f}%  Sharpe {m['sharpe_ratio']:.2f}")

    print(f"\nGenerating charts...")
    plot_equity_curves   (systems, output_dir / f'equity_curves_{ts}.png')
    plot_drawdowns       (systems, output_dir / f'drawdowns_{ts}.png')
    plot_monthly_returns (systems, output_dir / f'monthly_returns_{ts}.png')
    plot_pnl_distribution(systems, output_dir / f'pnl_distribution_{ts}.png')
    plot_annual_returns  (systems, output_dir / f'annual_returns_{ts}.png')
    plot_rolling_sharpe  (systems, output_dir / f'rolling_sharpe_{ts}.png')
    plot_trade_scatter   (systems, output_dir / f'trade_scatter_{ts}.png')
    plot_exit_reasons    (systems, output_dir / f'exit_reasons_{ts}.png')

    print(f"\nGenerating Excel report...")
    build_excel(systems, results_dir, output_dir / f'performance_report_{ts}.xlsx')

    # ── Console summary ───────────────────────────────────────────────────
    print(f"\n{'='*78}")
    print(f"{'System':<22} {'CAGR':>7} {'Alpha':>8} {'Sharpe':>7} "
          f"{'MaxDD':>7} {'WinRate':>8} {'PF':>5} {'Trades':>7}")
    print('-' * 78)
    print(f"{'SPY Buy & Hold':<22} {BENCHMARK_CAGR:>6.2f}%  {'—':>7}  {'—':>6}  "
          f"{'—':>6}  {'—':>7}  {'—':>4}  {'1':>6}")
    for sid, data in systems.items():
        m     = data['metrics']
        alpha = m['cagr'] - BENCHMARK_CAGR
        sign  = '+' if alpha >= 0 else ''
        print(f"{DISPLAY_NAMES[sid]:<22} {m['cagr']:>6.2f}%  "
              f"{sign}{alpha:>6.2f}%  {m['sharpe_ratio']:>6.2f}  "
              f"{m['max_drawdown_pct']:>6.2f}%  {m['win_rate']:>7.2f}%  "
              f"{m['profit_factor']:>4.2f}  {m['total_trades']:>7}")
    print('=' * 78)
    print(f"\nAll files written to: {output_dir}\n")


if __name__ == '__main__':
    main()
